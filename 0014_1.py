from openpyxl import Workbook
#from openpyxl.compat import range
from openpyxl.utils import get_column_letter
#openpyxl : https://openpyxl.readthedocs.io/en/default/usage.html
wb=Workbook()
filename=r"C:\Users\chauncey\Desktop\test.xlsx"
ws1=wb.active
for row in range(1,20):
    ws1.append(range(30))

ws2=wb.create_sheet(title="Pi")
ws2['C3']=3.14
ws2.cell(column=2,row=3,value=100)
ws3=wb.create_sheet(title="data")
for row in range(1,20):
    for col in range(3,10):
        _= ws3.cell(column=col,row=row,value=row+col)
print(ws2["C3"].value,ws2['B3'].value)
wb.save(filename)
