from multiprocessing import pool
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re
import requests
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
from collections import defaultdict,Iterable
def get_session(URL):
    response=Session.get(URL,headers=headers)
    bs=BeautifulSoup(response.content,"html.parser")
    return bs
def get_page_allLink(bs):
    for link in bs.find('ol',attrs={'class':"grid_view"}):
        try:
            L=link.find('a',href=re.compile("https://movie.douban.com/subject/.*$"))
            if str(L['href']) not in pages:
                pages.add(L['href'])
        except:
            pass

def get_movie_info(URL):
    Info=dict()
    bs=get_session(URL)
    try:
        Info["movie_name"] = bs.find("span",attrs={"property":"v:itemreviewed"}).get_text()
        Info["year"] = bs.find("span",attrs={"class":"year"}).get_text()
        Info["info"] = bs.find("div",{"id":"info"}).get_text().split('\n')
        while '' in Info['info']:Info['info'].remove('')
        Info['rating_num'] = bs.find("strong",attrs={'property':"v:average"}).get_text()
        Info['votes'] = bs.find("span",attrs={"property":"v:votes"}).get_text()
        Info["summary"] = bs.find("span",attrs={"property":"v:summary"}).get_text().strip('\n ').replace('\u3000','')
        Info["summary"] += bs.find("span",attrs={"class":"all hidden"}).get_text().strip('\n ').replace('\u3000','')
        Info["movie_url"]=str(URL)
    except:
        pass
    return Info
def format_xlsx(info,sheet,i):
    try:
       sheet.cell(row=1+i, column=1, value="电影名称");sheet.cell(row=1+i, column=2, value=info["movie_name"])
       sheet.cell(row=1+i, column=3, value="电影评分");sheet.cell(row=1+i, column=4, value=info["rating_num"])
       sheet.cell(row=1+i, column=5, value="投票人数");sheet.cell(row=1+i, column=6, value=info["votes"])
       sheet.cell(row=1+i, column=7, value="年份");sheet.cell(row=1+i, column=8, value=info["year"])
       for t in range(0,8):
        sheet.merge_cells(start_row=2+i+t, start_column=1, end_row=2+i+t, end_column=8)
       for t in range(2,12):
           sheet.cell(row=t + i, column=1, value=(info['info'][t-2]))
       sheet.merge_cells(start_row=11 + i , start_column=1, end_row=14 + i , end_column=8)
       sheet.cell(row=11+i,column=1,value=info["summary"].strip(' \n'))
       c=sheet.cell(row=1+i,column=2)
       c.font=Font(color=RED)
    except:
        pass
def save_execl(pages,xlsx_path):
    wb=Workbook()
    sheet=wb.active
    i=0
    while len(pages)>0:
        url=pages.pop()
        info=get_movie_info(url)
        format_xlsx(info,sheet,i)
        i+=15
    sheet.title="豆瓣top250"
    wb.save(xlsx_path)
if __name__=='__main__':
    URL = "https://movie.douban.com/top250"
    xlsx_path=r"C:\Users\chauncey\Desktop\豆瓣top250.xlsx"
    Session = requests.Session()
    pages = set()
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate, br',
        'Accept-Language': 'zh-CN,zh;q=0.8',
        'Connection': 'keep-alive',
        'Host': 'movie.douban.com',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.101 Safari/537.36'
    }
    for i in range(0,10):
        LURL=URL+'?start='+str(i*25)+'&filter='
        bs=get_session(LURL)
        get_page_allLink(bs)
    save_execl(pages,xlsx_path)









