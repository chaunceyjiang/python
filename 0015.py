import json,os
from openpyxl import Workbook
from collections import OrderedDict,Iterable
def json2xlsx(xlsx_path,json_path):
    wb=Workbook()
    ws1=wb.active
    ws1.title=xlsx_path.split('.')[0].split('\\')[-1]
    with open(json_path) as f:
        fulljson=f.read()
    j=json.loads(fulljson,object_pairs_hook=OrderedDict)#object_pairs_hook 保持json的顺序
    
    for row,row_val in enumerate(j):
        _ = ws1.cell(row=row+1,column=1,value=row_val)
        if isinstance(j[row_val],Iterable) and not isinstance(j[row_val],str):#可迭代对象且不是string类型
            for col,col_val in enumerate(j[row_val]):
                _= ws1.cell(row=row+1,column=col+2,value=col_val)
        else:
            _ = ws1.cell(row=row+1,column=2,value=j[row_val])
            
    wb.save(xlsx_path)

if __name__=='__main__':
    path=r'C:\Users\chauncey\Desktop'
    xlsx_name='City.xlsx'
    json_name='city.txt'
    json2xlsx(path+os.sep+xlsx_name,path+os.sep+json_name)
